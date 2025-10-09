"""This module contains the main process of the robot."""

from OpenOrchestrator.orchestrator_connection.connection import OrchestratorConnection
from OpenOrchestrator.database.queues import QueueElement
import os
import pandas as pd
import re
import xml.etree.ElementTree as ET
import requests
import json
from urllib.parse import quote
from datetime import datetime, timedelta
from office365.runtime.auth.user_credential import UserCredential
from office365.sharepoint.client_context import ClientContext
from office365.sharepoint.sharing.links.kind import SharingLinkKind
from office365.sharepoint.webs.web import Web
from office365.runtime.client_request_exception import ClientRequestException
from requests_ntlm import HttpNtlmAuth
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment, Font, Protection
import smtplib
from email.message import EmailMessage
from PIL import ImageFont, ImageDraw, Image
import pytz
import uuid
import GenerateNovaCase
import GetKmdAcessToken
import time
import pyodbc
# pylint: disable-next=unused-argument
def process(orchestrator_connection: OrchestratorConnection, queue_element: QueueElement | None = None) -> None:
    """Do the primary process of the robot."""
    orchestrator_connection.log_trace("Running process.")

    def shorten_document_title(doktitle):
        if len(doktitle) > 99:
            print(f'Dokumenttitlen {doktitle} er for lang {len(doktitle)}- afkortes')
            doktitle = doktitle[:95]
            return doktitle
        else:
            return doktitle
        
    log = True
    send_email = True

    #Getting credentials
    API_url = orchestrator_connection.get_constant("AktbobSharePointURL").value
    API_credentials = orchestrator_connection.get_credential("AktbobAPIKey")
    API_username = API_credentials.username
    API_password = API_credentials.password

    #Define developer mail
    UdviklerMail = orchestrator_connection.get_constant("balas").value

    #Get Robot Credentials
    RobotCredentials = orchestrator_connection.get_credential("Robot365User")
    RobotUsername = RobotCredentials.username
    RobotPassword = RobotCredentials.password

    queue_json = json.loads(queue_element.data)

    # Retrieve elements from queue_json
    SagsID = str(queue_json["SagsNummer"])
    MailModtager = str(queue_json["Email"])
    PodioID = str(queue_json["PodioID"])
    DeskProID = str(queue_json["DeskproID"])
    DeskProTitel = str(queue_json["Titel"])
    IndsenderNavn = str(queue_json['IndsenderNavn'])
    IndsenderMail = str(queue_json['IndsenderMail'])
    AktindsigtsDato = str(queue_json['AktindsigtsDato'])
    orchestrator_connection.log_info(f'Processing {SagsID} in {DeskProTitel}')
    AktSagsURL = str(queue_json['AktSagsURL'])
    memo_tunnel = False
    nul_dokument = False

    #Determining if it is a Nova-case or not
    pattern = r"^[A-Z]{3}-\d{4}-\d{6}"

    if re.match(pattern, SagsID):
        GeoSag = True
        NovaSag = False  
        GOAPILIVECRED = orchestrator_connection.get_credential("GOAktApiUser")
        GOAPILIVECRED_username = GOAPILIVECRED.username
        GOAPILIVECRED_password = GOAPILIVECRED.password
        GOAPI_URL = orchestrator_connection.get_constant('GOApiURL').value
    else:
        NovaSag = True
        GeoSag = False  

        def GetKMDToken(orchestrator_connection: OrchestratorConnection):
            KMD_Client = orchestrator_connection.get_credential("KMDClientSecret")
            client_secret = KMD_Client.password
            TokenTimeStamp = orchestrator_connection.get_constant("KMDTokenTimestamp").value
            KMD_access = orchestrator_connection.get_credential("KMDAccessToken")
            KMD_access_token = KMD_access.password
            KMD_URL = KMD_access.username
        
            # Define Danish timezone
            danish_timezone = pytz.timezone("Europe/Copenhagen")
        
            # Parse the old timestamp to a datetime object
            old_time = datetime.strptime(TokenTimeStamp.strip(), "%d-%m-%Y %H:%M:%S")
            old_time = danish_timezone.localize(old_time)  # Localize to Danish timezone
            print('Old timestamp: ' + old_time.strftime("%d-%m-%Y %H:%M:%S"))
        
            # Get the current timestamp in Danish timezone
            current_time = datetime.now(danish_timezone)
            print('current timestamp: '+current_time.strftime("%d-%m-%Y %H:%M:%S"))
            str_current_time = current_time.strftime("%d-%m-%Y %H:%M:%S")
        
            # Calculate the difference between the two timestamps
            time_difference = current_time - old_time
            print(time_difference)
        
            # Check if the difference is over 1 hour and 30 minutes
            GetNewTimeStamp = time_difference > timedelta(hours=1, minutes=30)
        
            # Output for the boolean
            print("GetNewTimeStamp:", GetNewTimeStamp)
        
            # Example of using it in an if-statement
            if GetNewTimeStamp:
                print("The difference is over 1 hour and 30 minutes. Fetch a new timestamp!")
                # Replace these values with your actual keys
                client_id = 'aarhus_kommune'
                client_secret = client_secret
                scope = 'client'
                grant_type = 'client_credentials'
        
        
                # Data to be sent in the POST request
                keys = {
                    'client_id': client_id,
                    'client_secret': client_secret,
                    'scope': scope,
                    'grant_type': grant_type,  # Specify the grant type you're using
                }
        
                # Sending POST request to get the access token
                response = requests.post(KMD_URL, data=keys)
        
                # Check if the request was successful (status code 200)
                if response.status_code == 200:
                    KMD_access_token = response.json().get('access_token')
                    print("Access token granted")
                    orchestrator_connection.update_credential("KMDAccessToken",KMD_URL,KMD_access_token)
                    orchestrator_connection.update_constant("KMDTokenTimestamp",str_current_time)
            
                    return KMD_access_token
                else:
                    print("Failed to get the access token")
        
            else:
                print("No need to fetch a new timestamp - using old timestamp")
                return KMD_access_token
        KMD_access_token = GetKMDToken(orchestrator_connection)   
        NOVA_URL = orchestrator_connection.get_constant("KMDNovaURL").value

    #Assigning different URL's depending on case type
    if NovaSag:
        Document_url = NOVA_URL + "/Document/GetList?api-version=2.0-Case"
        Case_url = NOVA_URL + "/Case/GetList?api-version=2.0-Case"
        id = str(uuid.uuid4())
    else:
        url = GOAPI_URL + "/_goapi/Cases/Metadata/" + SagsID

    if log:
        orchestrator_connection.log_info("Process starter")

    # Create session with NTLM authentication
    session = requests.Session()
    if GeoSag:
        session.auth = HttpNtlmAuth(GOAPILIVECRED_username, GOAPILIVECRED_password)
        session.headers.update({"Content-Type": "application/json"})
        response = session.get(url, timeout=500)
        response.raise_for_status()

        # Process the response content directly (assuming response.status_code == 200)
        SagMetaData = response.text
        json_obj = json.loads(SagMetaData)

        # Extract the "Metadata" field from the JSON response
        metadata_xml = json_obj.get("Metadata")
        if metadata_xml:
            # Parse the XML string
            xdoc = ET.fromstring(metadata_xml)

            # Extract attributes
            SagsURL = xdoc.attrib.get("ows_CaseUrl")
            SagsTitel = xdoc.attrib.get("ows_Title")

            # Process SagsURL
            if SagsURL and "cases/" in SagsURL:
                # Split SagsURL by "cases/" and take the second part
                Akt = SagsURL.split("cases/")[1].split("/")[0]
            else:
                print("Error: 'cases/' not found in SagsURL or SagsURL is missing.")
        else:
            print("Error: 'Metadata' field is missing in the JSON response.")
    if NovaSag:
        NewGuID = str(uuid.uuid4())
        payload = json.dumps({
        "common": {
            "transactionId": NewGuID
        },
        "paging": {
            "startRow": 1,
            "numberOfRows": 5
        },
        "CASEATTRIBUTES": {
            "USERFRIENDLYCASENUMBER": SagsID
        },
        "caseGetOutput": {
            "caseAttributes": {
            "title": True,
            "userFriendlyCaseNumber": True,
            "numberOfDocuments": True
            }
        }
        })
        headers = {
        "Authorization": f"Bearer {KMD_access_token}",
        'Content-Type': 'application/json'
        }

        response = requests.request("PUT", Case_url, headers=headers, data=payload)
        response.raise_for_status()

        # Process the response content directly (assuming response.status_code == 200)
        SagMetaData = response.text
        json_obj = json.loads(SagMetaData)

        SagsTitel = json_obj['cases'][0]['caseAttributes']['title']
        SagsURL = "" #SagsURL is nothing for now due to the setup in nova - potentially add later

    # Send GET request    
    if log:
        orchestrator_connection.log_info("Getting metadata")

    # Removal of illegal characters and double spaces
    pattern = r'[~#%&*{}\:\\<>?/+|\"\'\t\[\]`^@=!$();\€£¥₹]'
    SagsTitel = re.sub(pattern, '', str(SagsTitel))
    SagsTitel = " ".join(SagsTitel.split())
    document_without_date = False

    # Define the structure of the DataTable
    columns = [
        "Akt ID", "Dok ID", "Dokumenttitel", "Dokumentkategori", "Dokumentdato", 
        "Bilag til Dok ID", "Bilag", "Link til dokument", 
        "Omfattet af ansøgningen? (Ja/Nej)", "Gives der aktindsigt i dokumentet? (Ja/Nej/Delvis)", 
        "Begrundelse hvis nej eller delvis"
    ]

    # Create an empty DataFrame with these columns
    data_table = pd.DataFrame(columns=columns)

    if log:
        orchestrator_connection.log_info("Sagsurl" + SagsURL)
    if GeoSag:
        orchestrator_connection.log_info("Processing GEO case")
        Akt = SagsURL.split("/")[1]  
        
        if log:
            orchestrator_connection.log_info("Akt" + Akt)# Constructing the URL
        # Replacing '-' with '%2D' in SagsID
        encoded_sags_id = SagsID.replace("-", "%2D")
        ListURL = f"%27%2Fcases%2F{Akt}%2F{encoded_sags_id}%2FDokumenter%27"
        
        if log:
            orchestrator_connection.log_info("ListURL: " + ListURL)
        
        # Initialize variables
        ViewId = None
        view_ids_to_use = []  # To handle combined views
        response = session.get(f"{GOAPI_URL}/{SagsURL}/_goapi/Administration/GetLeftMenuCounter")
        ViewsIDArray = json.loads(response.text) # Parse the JSON

        # Check for "UdenMapper.aspx"
        for item in ViewsIDArray:
            if item["ViewName"] == "UdenMapper.aspx":
                ViewId = item["ViewId"]
                break
            elif item["ViewName"].lower() == "ikkejournaliseret.aspx":
                ikke_journaliseret_id = item["ViewId"]    
                if ikke_journaliseret_id is None: 
                    print('None detecteret')
                    LinkURL = item["LinkUrl"]
                    reponse = session.get(f'{GOAPI_URL}{LinkURL}')
                                    
                    # Find _spPageContextInfo JavaScript-objektet
                    match = re.search(r'_spPageContextInfo\s*=\s*({.*?});', reponse.text, re.DOTALL)
                    if not match:
                        raise ValueError("Kunne ikke finde _spPageContextInfo i HTML")
                    context_info = json.loads(match.group(1))
                    view_id = context_info.get("viewId", "").strip("{}")
                    ikke_journaliseret_id = view_id
            elif item["ViewName"] == "Journaliseret.aspx":
                journaliseret_id = item["ViewId"]
                if journaliseret_id is None:
                    LinkURL = item["LinkUrl"]
                    reponse = session.get(f'{GOAPI_URL}{LinkURL}')
                    match = re.search(r'_spPageContextInfo\s*=\s*({.*?});', reponse.text, re.DOTALL)
                    if not match:
                        raise ValueError("Kunne ikke finde _spPageContextInfo i HTML")
                    context_info = json.loads(match.group(1))
                    view_id = context_info.get("viewId", "").strip("{}")
                    journaliseret_id = view_id

        # If "UdenMapper.aspx" doesn't exist, combine views
        if ViewId is None:
            view_ids_to_use = [ikke_journaliseret_id, journaliseret_id]

        # Iterate through views
        for current_view_id in ([ViewId] if ViewId else view_ids_to_use):
            firstrun = True
            MorePages = True

            while MorePages:
                if log:
                    orchestrator_connection.log_info("Henter dokumentlister")

                # If not the first run, fetch the next page
                if not firstrun:
                    url = f"{GOAPI_URL}/{SagsURL}/_api/web/GetList(@listUrl)/RenderListDataAsStream"
                    url_with_query = f"{url}?@listUrl={ListURL}{NextHref.replace('?', '&')}"

                    response = session.post(url_with_query, timeout=500)
                    response.raise_for_status()
                    Dokumentliste = response.text  # Extract the content
                else:
                    # If first run, fetch the first page for the current view
                    url = f"{GOAPI_URL}/{SagsURL}/_api/web/GetList(@listUrl)/RenderListDataAsStream"
                    query_params = f"?@listUrl={ListURL}&View={current_view_id}"
                    full_url = url + query_params

                    response = session.post(full_url, timeout=500)
                    response.raise_for_status()
                    Dokumentliste = response.text  # Extract the content

                # Deserialize response
                dokumentliste_json = json.loads(Dokumentliste)
                dokumentliste_rows = dokumentliste_json.get("Row", [])

                # Check for additional pages
                NextHref = dokumentliste_json.get("NextHref")
                MorePages = "NextHref" in dokumentliste_json

                # Process each row
                for item in dokumentliste_rows:
                    # Extract and prepare data
                    DokumentURL = GOAPI_URL.replace("ad.", "") + quote(item.get("FileRef", ""), safe="/")
                    AktID = item.get("CaseRecordNumber", "").replace(".", "")
                    DokumentDato = str(item.get("Dato"))
                    if not DokumentDato:
                        document_without_date = True
                    Dokumenttitel = item.get("Title", "")
                    DokID = str(item.get("DocID"))
                    DokumentKategori = str(item.get("Korrespondance"))

                    if len(Dokumenttitel) < 2:
                        Dokumenttitel = item.get("FileLeafRef.Name", "")
                    if str(AktID).strip() == '0':
                        orchestrator_connection.log_info('0 dokument detekteret')
                        nul_dokument = True

                    if log:
                        orchestrator_connection.log_info(f"AktID: {AktID}")

                    # Fetch parents and children data
                    parents_response = session.get(f"{GOAPI_URL}/_goapi/Documents/Parents/{DokID}", timeout=500)
                    parents_object = json.loads(parents_response.text)
                    ParentArray = parents_object.get("ParentsData", [])
                    Bilag = ", ".join(str(currentItem.get("DocumentId", "")) for currentItem in ParentArray)

                    children_response = session.get(f"{GOAPI_URL}/_goapi/Documents/Children/{DokID}", timeout=500)
                    children_object = json.loads(children_response.text)
                    ChildrenArray = children_object.get("ChildrenData", [])
                    BilagChild = ", ".join(str(currentItem.get("DocumentId", "")) for currentItem in ChildrenArray)

                    # Append data to DataFrame
                    if "tunnel_marking" in Dokumenttitel.lower() or "memometadata" in Dokumenttitel.lower():
                        memo_tunnel = True
                        data_table = pd.concat([data_table, pd.DataFrame([{
                            "Akt ID": AktID,
                            "Dok ID": DokID,
                            "Dokumenttitel": shorten_document_title(Dokumenttitel),
                            "Dokumentkategori": DokumentKategori,
                            "Dokumentdato": DokumentDato,
                            "Bilag": BilagChild,
                            "Bilag til Dok ID": Bilag,
                            "Link til dokument": DokumentURL,
                            "Omfattet af ansøgningen? (Ja/Nej)": "Ja",
                            "Gives der aktindsigt i dokumentet? (Ja/Nej/Delvis)": "Nej",
                            "Begrundelse hvis nej eller delvis": "Tavshedsbelagte oplysninger - om private forhold"
                        }])], ignore_index=True)
                    else:
                        data_table = pd.concat([data_table, pd.DataFrame([{
                            "Akt ID": AktID,
                            "Dok ID": DokID,
                            "Dokumenttitel": shorten_document_title(Dokumenttitel),
                            "Dokumentkategori": DokumentKategori,
                            "Dokumentdato": DokumentDato,
                            "Bilag": BilagChild,
                            "Bilag til Dok ID": Bilag,
                            "Link til dokument": DokumentURL,
                            "Omfattet af ansøgningen? (Ja/Nej)": "Ja",
                            "Gives der aktindsigt i dokumentet? (Ja/Nej/Delvis)": "",
                            "Begrundelse hvis nej eller delvis": ""
                        }])], ignore_index=True)

                firstrun = False
    else:
        orchestrator_connection.log_info("Processing NOVA case")
        payload = json.dumps({
        "common": {
            "transactionId": id
        },
        "paging": {
            "startRow": 0,
            "numberOfRows": 10000,
            "calculateTotalNumberOfRows": True
        },
        "caseNumber": SagsID,
        "getOutput": {
            "documentType": True,
            "title": True,
            "caseWorker": True,
            "description": True, 
            "fileExtension": True,
            "approved": True, 
            "acceptReceived": True,
            "documentDate": True
        }
        })
        headers = {
        "Authorization": f"Bearer {KMD_access_token}",
        'Content-Type': 'application/json'
        }

        response = requests.request("PUT", Document_url, headers=headers, data=payload)
        response.raise_for_status
        aktid_number = 1
        documents = json.loads(response.text)['documents']
        # Process each row
        for i in range(len(documents)):
            # Extract and prepare data
            DokumentURL = ""
            AktID = aktid_number
            DokumentDato = str(documents[i]['documentDate'])
            if not DokumentDato:
                document_without_date = True
            date_object = datetime.strptime(DokumentDato, "%Y-%m-%dT%H:%M:%S")
            formatted_date = str(date_object.strftime("%d-%m-%Y"))
            Dokumenttitel = documents[i]['title']
            DokID = documents[i]['documentNumber']
            DokumentKategori = documents[i]['documentType']

            # Append data to DataFrame
            if "tunnel_marking" in Dokumenttitel.lower() or "memometadata" in Dokumenttitel.lower() or "fletteliste" in Dokumenttitel.lower():
                memo_tunnel = True
                data_table = pd.concat([data_table, pd.DataFrame([{
                    "Akt ID": AktID,
                    "Dok ID": DokID,
                    "Dokumenttitel": shorten_document_title(Dokumenttitel),
                    "Dokumentkategori": DokumentKategori,
                    "Dokumentdato": formatted_date,
                    "Bilag": "",
                    "Bilag til Dok ID": "",
                    "Link til dokument": DokumentURL,
                    "Omfattet af ansøgningen? (Ja/Nej)": "Ja",
                    "Gives der aktindsigt i dokumentet? (Ja/Nej/Delvis)": "Nej",
                    "Begrundelse hvis nej eller delvis": "Tavshedsbelagte oplysninger - om private forhold"
                }])], ignore_index=True)
            else:
                data_table = pd.concat([data_table, pd.DataFrame([{
                    "Akt ID": AktID,
                    "Dok ID": DokID,
                    "Dokumenttitel": shorten_document_title(Dokumenttitel),
                    "Dokumentkategori": DokumentKategori,
                    "Dokumentdato": formatted_date,
                    "Bilag": "",
                    "Bilag til Dok ID": "",
                    "Link til dokument": DokumentURL,
                    "Omfattet af ansøgningen? (Ja/Nej)": "Ja",
                    "Gives der aktindsigt i dokumentet? (Ja/Nej/Delvis)": "",
                    "Begrundelse hvis nej eller delvis": ""
                }])], ignore_index=True)
            aktid_number += 1

    SMTP_SERVER = "smtp.adm.aarhuskommune.dk"
    SMTP_PORT = 25
    SCREENSHOT_SENDER = "aktbob@aarhus.dk"

    if document_without_date:
        send_missing_documentdate(MailModtager, SagsID, SCREENSHOT_SENDER, UdviklerMail, SMTP_SERVER, SMTP_PORT)
        return

    # Define font settings
    FONT_PATH = "calibri.ttf"  # Ensure this file exists in your directory
    FONT_SIZE = 11

    # Load the font
    try:
        font = ImageFont.truetype(FONT_PATH, FONT_SIZE)
    except OSError:
        raise FileNotFoundError(f"Font file not found at {FONT_PATH}. Please ensure the font file is available.")

    # Function to calculate text dimensions in Excel units
    def calculate_text_dimensions(text, font, max_width_in_pixels):
        dummy_image = Image.new("RGB", (1, 1))
        draw = ImageDraw.Draw(dummy_image)
        bbox = draw.textbbox((0, 0), text, font=font)
        text_width = bbox[2] - bbox[0]
        text_height = bbox[3] - bbox[1]
        excel_column_width = text_width / 5
        lines = max(1, text_width // max_width_in_pixels + 1)
        excel_row_height = lines * (text_height / 1.33)
        return excel_column_width, excel_row_height
    
    tom_sag = False

    if data_table.empty:
        tom_sag = True
        fake_row = {col: "" for col in data_table.columns}
        data_table = pd.DataFrame([fake_row])  # Add placeholder row

    # Ensure 'Akt ID' is numeric and clean
    data_table['Akt ID'] = pd.to_numeric(data_table['Akt ID'].astype(str).str.strip(), errors='coerce')

    # Sort values if the table is not empty
    if not data_table.empty:
        data_table = data_table.sort_values(by='Akt ID', ascending=True, ignore_index=True)

    # 🟢 STEP 2: SAVE THE DATAFRAME TO EXCEL
    excel_file_path = f"{SagsID}_{datetime.now().strftime('%d-%m-%Y')}.xlsx"
    data_table.to_excel(excel_file_path, index=False, sheet_name="Sagsoversigt")

    # Open Excel file for formatting
    workbook = load_workbook(excel_file_path)
    worksheet = workbook["Sagsoversigt"]

    # Ensure at least 2 rows exist (header + data row)
    if worksheet.max_row == 1:
        worksheet.append([""] * worksheet.max_column)  # Add an empty row

    data_range = f"A1:K{worksheet.max_row}"
    table = Table(displayName="SagsoversigtTable", ref=data_range)
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
                        showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    worksheet.add_table(table)

    # Apply column width formatting dynamically
    max_width_in_pixels = 382
    for col_idx, column_cells in enumerate(worksheet.columns, start=1):
        max_length = 0
        for cell in column_cells:
            if cell.value:
                text = str(cell.value)
                column_width, _ = calculate_text_dimensions(text, font, max_width_in_pixels)
                max_length = max(max_length, column_width)
        worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 4, 50)

    # Specific column adjustments
    COLUMN_C_INDEX, COLUMN_G_INDEX = 3, 7
    worksheet.column_dimensions[get_column_letter(COLUMN_C_INDEX)].width = 50

    # Define header styling
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    for cell in worksheet[1]:
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.font = header_font
    
    # Apply row height adjustments for wrapped text
    ROW_HEIGHT_PER_PIXEL = 1
    def calculate_row_height(text, font, max_width_in_pixels):
        dummy_image = Image.new("RGB", (1, 1))
        draw = ImageDraw.Draw(dummy_image)
        bbox = draw.textbbox((0, 0), text, font=font)
        text_width = bbox[2] - bbox[0]
        text_height = bbox[3] - bbox[1]
        lines_required = max(1, (text_width / max_width_in_pixels) + 1)
        return lines_required * text_height * ROW_HEIGHT_PER_PIXEL

    # Adjust row heights for Columns C and G
    for row_idx in range(2, worksheet.max_row + 1):
        row_height = 15
        for col_idx in [COLUMN_C_INDEX, COLUMN_G_INDEX]:
            cell = worksheet.cell(row=row_idx, column=col_idx)
            if cell.value:
                cell.alignment = Alignment(wrap_text=True)
                text = str(cell.value)
                height = calculate_row_height(text, font, 150 if col_idx == COLUMN_C_INDEX else 70)
                row_height = max(row_height, height)
        worksheet.row_dimensions[row_idx].height = row_height
    
    for col in ["I", "J", "K"]:
        for row_idx in range(2, worksheet.max_row + 1):  # Skip header
            cell = worksheet[f"{col}{row_idx}"]
            cell.protection = Protection(locked=False)  # Allow dropdown selection

    # Add hyperlinks in column H
    for row_idx in range(2, worksheet.max_row + 1):
        cell = worksheet.cell(row=row_idx, column=8)
        if cell.value:
            cell.value, cell.hyperlink, cell.style = "Dokumentlink", cell.value, "Hyperlink"

    # Add dropdown validations
    validation_i = DataValidation(type="list", formula1='"Ja,Nej"', allow_blank=False, showErrorMessage=True)
    validation_i.error, validation_i.errorTitle = "Vælg venligst Ja eller Nej.", "Ugyldig værdi"

    validation_j = DataValidation(type="list", formula1='"Ja,Delvis,Nej"', allow_blank=False, showErrorMessage=True)
    validation_j.error, validation_j.errorTitle = "Vælg venligst Ja, Delvis eller Nej.", "Ugyldig værdi"

    # Create hidden sheet for dropdown options
    hidden_options = [
        "Internt dokument - ufærdigt arbejdsdokument",
        "Internt dokument - foreløbige og sagsforberedende overvejelser",
        "Internt dokument - del af intern beslutningsproces",
        "Særlige dokumenter - korrespondance med sagkyndig rådgiver vedr. tvistsag",
        "Særlige dokumenter - statistik og undersøgelser",
        "Særlige dokumenter - straffesag",
        "Tavshedsbelagte oplysninger - om private forhold",
        "Tavshedsbelagte oplysninger - forretningsforhold",
        "Tavshedsbelagte oplysninger - Andet (uddybes i afgørelsen)",
        " "
    ]

    hidden_sheet = workbook.create_sheet("VeryHidden")
    hidden_sheet.sheet_state = "veryHidden"
    for idx, option in enumerate(hidden_options, start=1):
        hidden_sheet.cell(row=idx, column=1, value=option)

    # Add validation for column K using hidden sheet values
    validation_k = DataValidation(type="list", formula1=f"=VeryHidden!$A$1:$A${len(hidden_options)}",
                                allow_blank=False, showErrorMessage=True)
    validation_k.error, validation_k.errorTitle = "Vælg en mulighed.", "Ugyldig indtastning"

    first_data_row = 2 if worksheet.max_row > 1 else 1
    validation_i.add(f"I{first_data_row}:I{worksheet.max_row}")
    validation_j.add(f"J{first_data_row}:J{worksheet.max_row}")
    validation_k.add(f"K{first_data_row}:K{worksheet.max_row}")

    worksheet.add_data_validation(validation_i)
    worksheet.add_data_validation(validation_j)
    worksheet.add_data_validation(validation_k)

    worksheet.protection.sheet = True
    worksheet.protection.password = "Aktbob"
    worksheet.protection.enable()

    workbook.save(excel_file_path)

    Mappe1 = str(DeskProID) +" - " + str(DeskProTitel)
    Mappe2 = str(SagsID) + " - " + str(SagsTitel)

    # Authenticate to SharePoint using Office365 credentials
    credentials = UserCredential(RobotUsername, RobotPassword)
    ctx = ClientContext(API_url).with_credentials(credentials)

    certification = orchestrator_connection.get_credential("SharePointCert")
    api = orchestrator_connection.get_credential("SharePointAPI")

    cert_credentials = {
        "tenant": api.username,
        "client_id": api.password,
        "thumbprint": certification.username,
        "cert_path": certification.password
    }

    ctx = ClientContext(API_url).with_client_certificate(**cert_credentials)

    # Function to sanitize folder names
    def sanitize_folder_name(folder_name):
        pattern = r'[.,~#%&*{}\[\]\\:<>?/+|$¤£€\"\t]'
        folder_name = re.sub(pattern, "", folder_name)
        folder_name = re.sub(r"\s+", " ", folder_name).strip()
        return folder_name

    # Sanitize folder names
    Mappe1 = sanitize_folder_name(Mappe1)
    Mappe2 = sanitize_folder_name(Mappe2)

    # Ensure folder names don't exceed length limits
    if len(Mappe1) > 99:
        Mappe1 = Mappe1[:95] + "(...)"
    if len(Mappe2) > 99:
        Mappe2 = Mappe2[:95] + "(...)"

    total_length = len(Mappe1) + len(Mappe2) + 17  # 17 is for folder structure overhead
    if total_length > 290:
        excess_length = total_length - 290
        half_excess = excess_length // 2
        Mappe1 = Mappe1[:len(Mappe1) - half_excess - 5] + "(...)"
        Mappe2 = Mappe2[:len(Mappe2) - half_excess - 5] + "(...)"

    parent_folder_name = API_url.split(".com")[-1] + "/Delte dokumenter/Dokumentlister" 

    # Create main folder
    root_folder = ctx.web.get_folder_by_server_relative_url(parent_folder_name)
    main_folder = root_folder.folders.add(Mappe1) 
    ctx.execute_query()

    # Create subfolder inside main folder
    subfolder = main_folder.folders.add(Mappe2)
    ctx.execute_query()

    file_path = excel_file_path  # Ensure it points to the created Excel file



    # Check if the file exists and upload it
    try:
        with open(file_path, "rb") as file_content:
            subfolder.upload_file(os.path.basename(file_path), file_content.read())
        ctx.execute_query()
    except ClientRequestException as e:
        if e.response is not None and e.response.status_code == 423:
            orchestrator_connection.log_info("File is locked (HTTP 423 Locked).")
            error_json = json.loads(e.response.text)
            error_code = error_json["error"]["code"]
            error_message = error_json["error"]["message"]["value"]
            send_dokumentliste_locked(MailModtager, SagsID, SCREENSHOT_SENDER, UdviklerMail, SMTP_SERVER, SMTP_PORT, error_code, error_message)
            return
        else:
            raise

    if log:
        orchestrator_connection.log_info("Folders created in sharepoint")

    # Step 2: Create a sharing link (e.g., Anonymous View Link)
    result = subfolder.share_link(SharingLinkKind.OrganizationEdit).execute_query()
    link_url = result.value.sharingLinkInfo.Url

    # Step 3: Verify the sharing link
    result = Web.get_sharing_link_kind(ctx, link_url).execute_query()


    memodata_obs = (
        "Vær opmærksom på, at denne sag indeholder dokumenter af typen memometadata, tunnel-marking eller fra flettelisten. Disse er automatisk sat til 'Nej', da de kan indeholde fortrolige oplysninger. Er dette forkert, kan du blot sætte dem til 'Ja' eller 'Delvis'."
        if memo_tunnel else ""
    )

    nuldokument_obs = (
        "Vær opmærksom på, at denne sag indeholder dokumenter, der er nul-dokumenter."
        if nul_dokument else ""
    )
    orchestrator_connection.log_info(f'{memodata_obs} {nuldokument_obs}')

    body = f"""
    <html>
        <body>
            <p>Sag: {DeskProID} - {DeskProTitel}. </p>
            <p>Der er bedt om aktindsigt i sag {SagsID}. Der er dannet en dokumentliste for sagen.</p>
            <a href="{link_url}">Link til dokumentlisten</a>

            <p>{memodata_obs}</p>
            <p>{nuldokument_obs}</p>

            <ul>
            <li>
                Kolonnen ”Omfattet af ansøgning”: 
                Dokumentlisten omfatter alle dokumenterne på sagen. Marker med ”nej”, hvis der er dokumenter, der ikke er omfattet af ansøgningen. (Styrer sammen med næste kolonne, hvilke dokumenter der sendes til ansøger.)
            </li>
            <li>
                Kolonnen ”Gives der aktindsigt i dokumentet?”: Marker for de dokumenter, der er omfattet af ansøgningen, om de skal udleveres. (Styrer sammen med foregående kolonne, hvilke dokumenter der sendes til ansøger.)
            </li>
            <li>
                Kolonnen ”Begrundelse hvis nej eller delvis”: Marker for de dokumenter, som ikke skal udleveres, hvilken type af undtagelse du bruger. (Styrer, hvilke begrundelsesfraser der indsættes i afgørelsen.)
            </li>
            </ul>
            <p>
            Markeringer i dokumentlisten kan ændres efterfølgende, hvis det er nødvendigt, så længe aktindsigten ikke er sendt.
            </p>
            <p>Er du i tvivl, om noget kan undtages, kan du finde hjælp her: <a href= "https://aktindsigtshaandbogen.dk/aktindsigt-trin-for-trin/?KeyListPageQuery=5404" target="_blank">Guide til undtagelser</a> under "Undtagelser fra aktindsigt" </p>
            <p>De øvrige kolonner kan ikke redigeres. Er der fortrolige oplysninger i selve aktlisten, skal disse overstreges på anden vis, inden aktindsigten sendes. </p>
            <p>Når dokumentlisten er gennemgået, kan screeningen af de valgte filer sættes i gang i dokumentstyringssystemet (Podio). </p>
            
            <br><br>
            <p> Vejledning findes på <a href="https://aarhuskommune.atlassian.net/wiki/spaces/AB/pages/64979049/AKTBOB+--+Vejledning" target="_blank">AKTBOB – Vejledning</a> </p>
            <p> Link til sagsstyringssystemet <a href="https://mtmsager.aarhuskommune.dk/app#/t/ticket/{DeskProID}" target="_blank">Sagsstyringssystem</a> </p>
            

        </body>
    </html>
    """


    def send_success_email(to_address: str | list[str], sags_id: str, body):
        """
        Sends an email notification with the provided body and subject.
        Args:
            to_address (str | list[str]): Email address or list of addresses to send the notification.
            sags_id (str): The ID of the case (SagsID) used in the email subject.
            deskpro_id (str): The DeskPro ID for constructing the DeskPro link.
            sharepoint_link (str): The SharePoint link to include in the email body.
        """
        # Email subject
        subject = f"{sags_id}: Dokumentliste oprettet"
        # Create the email message
        msg = EmailMessage()
        msg['To'] = ', '.join(to_address) if isinstance(to_address, list) else to_address
        msg['From'] = SCREENSHOT_SENDER
        msg['Subject'] = subject
        msg.set_content("Please enable HTML to view this message.")
        msg.add_alternative(body, subtype='html')
        msg['Reply-To'] = UdviklerMail
        msg['Bcc'] = UdviklerMail

        # Send the email using SMTP
        try:
            with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as smtp:
                smtp.send_message(msg)
                
        except Exception as e:
                print(f"Failed to send success email: {e}")
    
    def send_sag_empty_email(to_address: str | list[str], sags_id: str):

        # Email subject
        subject = f"{sags_id} er en tom sag"

        # Email body (HTML)
        body = f"""
        <html>
            <body>
                <p>Sagen {sags_id} er en tom sag. Vær opmærksom på, at processen ikke kan behandle tomme sager.</p>
            </body>
        </html>
        """
        # Create the email message
        msg = EmailMessage()
        msg['To'] = ', '.join(to_address) if isinstance(to_address, list) else to_address
        msg['From'] = SCREENSHOT_SENDER
        msg['Subject'] = subject
        msg.set_content("Please enable HTML to view this message.")
        msg.add_alternative(body, subtype='html')
        msg['Reply-To'] = UdviklerMail
        msg['Bcc'] = UdviklerMail

        # Send the email using SMTP
        try:
            with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as smtp:
                smtp.send_message(msg)
                
        except Exception as e:
            print(f"Failed to send success email: {e}")

    if log:
        orchestrator_connection.log_info("Sending email")

    # Encode folder names for URL safety
    Mappe1_encoded = quote(Mappe1)
    Mappe2_encoded = quote(Mappe2)

    # Construct the full SharePoint URL
    SharepointLink = f"{API_url}/Delte%20dokumenter/Dokumentlister/{Mappe1_encoded}/{Mappe2_encoded}"

    if send_email and tom_sag is not True:
        send_success_email(MailModtager, SagsID, body)
    if send_email and tom_sag is True:
        send_sag_empty_email(MailModtager, SagsID)
        orchestrator_connection.log_info('Email sent of empty case')

    if log:
        orchestrator_connection.log_info("Tilføjer link til Podio")

    # API Headers
    headers = {
        "ApiKey": API_password,  
        "Content-Type": "application/json"
    }

    # 1. PUT Request to Update Podio Link
    put_url = f"{API_username}/Podio/{PodioID}/DokumentlisteField"
    put_body = {"value": SharepointLink}
    put_response = requests.put(put_url, headers=headers, json=put_body)

    if put_response.status_code != 200 and put_response.status_code != 204:
        print(f"PUT request failed: {put_response.status_code}, {put_response.text}")

    # Debugging: Print URLs and Headers
    get_ticket_url = f"{API_username}/tickets?deskproId={quote(DeskProID)}"
    get_case_url = f"{API_username}/cases?podioItemId={quote(PodioID)}"

    # 2. GET Request to Fetch Ticket ID based on DeskProId
    get_ticket_response = requests.get(get_ticket_url, headers=headers, json={})  # Added json={}
    get_ticket_response.raise_for_status()

    ticket_data = get_ticket_response.json()
    ticket_id = ticket_data[0]['id']
    patch_ticket_url = f"{API_username}/tickets/{ticket_id}"
    patch_ticket_body = {"sharepointFolderName": Mappe1}
    patch_ticket_response = requests.patch(patch_ticket_url, headers=headers, json=patch_ticket_body)
    patch_ticket_response.raise_for_status()

    # 4. GET Request to Fetch Case ID based on PodioID
    get_case_response = requests.get(get_case_url, headers=headers, json={}) 
    get_case_response.raise_for_status()
    case_data = get_case_response.json()
    case_id = case_data[0]['id']
                
    patch_case_url = f"{API_username}/cases/{case_id}"
    patch_case_body = {"sharepointFolderName": Mappe2}
    patch_case_response = requests.patch(patch_case_url, headers=headers, json=patch_case_body)

    patch_case_response.raise_for_status()

    if os.path.exists(excel_file_path):
        os.remove(excel_file_path)

    def try_register_case(deskpro_id, lock_expiry_minutes=3):
        """
        Attempts to acquire a lock for the given DeskProID by inserting a row.
        Returns True if lock acquired, False if already held by another process.
        """
        now = datetime.utcnow()
        conn = pyodbc.connect(
            "DRIVER={ODBC Driver 17 for SQL Server};"
            "SERVER=srvsql29;"
            "DATABASE=PyOrchestrator;"
            "Trusted_Connection=yes"
        )
        cursor = conn.cursor()
        try:
            # Clean up old locks older than N minutes
            cursor.execute(
                """
                DELETE FROM dbo.NovaCaseRegistry
                WHERE CreatedAt < DATEADD(MINUTE, -?, GETUTCDATE())
                """,
                lock_expiry_minutes
            )

            try:
                # Try to insert a lock row for this DeskProID
                cursor.execute(
                    """
                    INSERT INTO dbo.NovaCaseRegistry (DeskProID, CreatedAt)
                    VALUES (?, ?)
                    """,
                    deskpro_id,
                    now
                )
                conn.commit()
                return True

            except pyodbc.IntegrityError:
                # Lock already held for this DeskProID
                return False

        finally:
            cursor.close()
            conn.close()


    def register_case_with_retry(deskpro_id, max_retries=10, delay_seconds=30):
        """
        Attempts to acquire a lock with retries.
        """
        for attempt in range(1, max_retries + 1):
            if try_register_case(deskpro_id):
                print(f"[Attempt {attempt}] Lock acquired for DeskProID {deskpro_id} — proceeding.")
                return True
            else:
                print(f"[Attempt {attempt}] Lock for DeskProID {deskpro_id} is held. Retrying in {delay_seconds} seconds...")
                time.sleep(delay_seconds)

        print(f"Max retries reached — aborting case creation for DeskProID {deskpro_id}.")
        return False
    

    if NovaSag and register_case_with_retry(DeskProID):
        KMD_access_token = GetKmdAcessToken.GetKMDToken(orchestrator_connection=orchestrator_connection)
        GenerateNovaCase.invoke_GenerateNovaCase(Sagsnummer = SagsID, KMDNovaURL= NOVA_URL, KMD_access_token = KMD_access_token, AktSagsURL= AktSagsURL, IndsenderNavn = IndsenderNavn, IndsenderMail= IndsenderMail, AktindsigtsDato = AktindsigtsDato, orchestrator_connection= orchestrator_connection, DeskProID = DeskProID)
    elif NovaSag:
        orchestrator_connection.log_info("Skipping Nova case generation — already in progress or created by another process.")

def send_dokumentliste_locked(to_address: str | list[str], sags_id: str, SCREENSHOT_SENDER, UdviklerMail, SMTP_SERVER, SMTP_PORT, error_code, error_message):

    # Email subject
    subject = f"Dokumentliste for {sags_id} er låst"

    # Email body (HTML)
    body = f"""
    <html>
        <body>
            <p>Dokumentlisten for {sags_id} er låst, og derfor kan robotten ikke generere en ny dokumentliste. Sørg for at lukke dokumentlisten ned på alle computere der kan have den åben, både i browseren og excel, og prøv at generere dokumentlisten igen.</p>
            <br>
            <p><b>Fejlinfo:</b></p>
        <ul>
            <li><b>Kode:</b> {error_code}</li>
            <li><b>Besked:</b> {error_message}</li>
        </ul>
        </body>
    </html>
    """
    # Create the email message
    msg = EmailMessage()
    msg['To'] = ', '.join(to_address) if isinstance(to_address, list) else to_address
    msg['From'] = SCREENSHOT_SENDER
    msg['Subject'] = subject
    msg.set_content("Please enable HTML to view this message.")
    msg.add_alternative(body, subtype='html')
    msg['Reply-To'] = UdviklerMail
    msg['Bcc'] = UdviklerMail

    # Send the email using SMTP
    try:
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as smtp:
            smtp.send_message(msg)
            
    except Exception as e:
        print(f"Failed to send locked email: {e}")

def send_missing_documentdate(to_address: str | list[str], sags_id: str, SCREENSHOT_SENDER, UdviklerMail, SMTP_SERVER, SMTP_PORT):

    # Email subject
    subject = f"{sags_id} indeholder dokumenter uden dato"

    # Email body (HTML)
    body = f"""
    <html>
        <body>
            <p>{sags_id} indeholder dokumenter i GO, der mangler dato. Sørg for at alle dokumenter i originalsagen har en dato, og genkør derefter dokumentlisten.</p>
        </body>
    </html>
    """
    # Create the email message
    msg = EmailMessage()
    msg['To'] = ', '.join(to_address) if isinstance(to_address, list) else to_address
    msg['From'] = SCREENSHOT_SENDER
    msg['Subject'] = subject
    msg.set_content("Please enable HTML to view this message.")
    msg.add_alternative(body, subtype='html')
    msg['Reply-To'] = UdviklerMail
    msg['Bcc'] = UdviklerMail

    # Send the email using SMTP
    try:
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as smtp:
            smtp.send_message(msg)
            
    except Exception as e:
        print(f"Failed to send locked email: {e}")
        